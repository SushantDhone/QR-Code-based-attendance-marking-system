'''Requirements:
pip install opencv-python
pip install pyzbar
pip install tk / pip install tkinter
pip install openpyxl'''

import cv2
import pyzbar.pyzbar as pyzbar
import tkinter as tk
from tkinter import PhotoImage
import openpyxl

# Create a set to store scanned QR codes
scanned_codes = set()

# File for attendance
attendance_file = 'attendance.xlsx'

# Function to update attendance
def update_attendance(code):
    try:
        wb = openpyxl.load_workbook(attendance_file)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.save(attendance_file)

    sheet = wb.active

    # Check if the code is already in the spreadsheet
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        if row[0] == code:
            break
    else:
        # Add code to the spreadsheet
        sheet.append([code])

    wb.save(attendance_file)
    wb.close()

# Initialize the webcam
cap = cv2.VideoCapture(0)

# Create a Tkinter window
window = tk.Tk()
window.title("QR Code Attendance System")

# Create a Canvas for displaying the video feed
canvas = tk.Canvas(window, width=640, height=480)
canvas.pack()

# Function to process video frames
def process_frame():
    ret, frame = cap.read()
    if ret:
        # Decode QR codes in the frame
        decoded_objects = pyzbar.decode(frame)
        
        for obj in decoded_objects:
            qr_data = obj.data.decode('utf-8')
            
            if qr_data not in scanned_codes:
                scanned_codes.add(qr_data)
                print(f"Scanned QR Code: {qr_data}")
                update_attendance(qr_data)

        # Convert BGR image to RGB format for Tkinter
        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

        # Create a PhotoImage from the frame
        photo = PhotoImage(data=cv2.imencode('.ppm', frame_rgb)[1].tobytes())
        canvas.create_image(0, 0, image=photo, anchor=tk.NW)
        canvas.photo = photo  # To prevent garbage collection

    window.after(10, process_frame)

# Start processing frames
process_frame()

# Run the Tkinter main loop
window.mainloop()

# Release the camera
cap.release()